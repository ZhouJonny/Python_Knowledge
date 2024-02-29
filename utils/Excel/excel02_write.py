import random

import openpyxl


"""第一步：创建工作簿（Workbook）"""
wb = openpyxl.Workbook()

"""第二步：添加工作表（Worksheet）"""
# sheet = wb.create_sheet("期末成绩")  # 会自动创建一个工作簿sheet，此代码会创建一个新的工作表
sheet = wb.active
sheet.title = '期末成绩'

"""第三步：向工作表中写入数据"""
titles = ('姓名', '语文', '数学', '英语')
# todo:创建表格第一行
for col_index, title in enumerate(titles):  # col_index从0开始，所以col_index+1
    sheet.cell(1, col_index+1, title)  # 在第一行，从第一列到第四列写入titles

names = ('关羽', '张飞', '赵云', '马超', '黄忠')
# todo:写入5个学生3门课程的成绩，成绩用50-100的随机数表示
for row_index, name in enumerate(names):
    sheet.cell(row_index+2, 1, name)  # 从第二行开始，在第一列写入names
    for col_index in range(2, 5):
        sheet.cell(row_index+2, col_index, random.randrange(50, 101))

"""第四步：保存工作簿"""
wb.save('Data/score.xlsx')