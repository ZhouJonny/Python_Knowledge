"""
python操作Excel

三方库：
    1。xlwt/xlrd/xlutils ---> 兼容低版本的Excel文件
    2。openpyxl
        注意 cell的行列索引都是从1开始的。
"""
# import xlrd
import openpyxl

# # 工作簿 ---> 一个Excel文件 ---> Workbook
# wb = xlrd.open_workbook('Data/info.xlsx')
# print(wb.sheet_names())  # 只适用于xls文件

wb = openpyxl.load_workbook('Data/info.xlsx')
print(type(wb))
# 获取所有工作表的名字
print(wb.sheetnames)
# 获取工作表
# sheet = wb['Worksheet']  # 这种写法会让pycharm不知道sheet类型，从而使用sheet对象时没有代码提示
sheet = wb.worksheets[0]  # 强烈推荐这样写，sheet会有代码提示
# A1:N24 工作表的纬度，左上角到右下角
print(sheet.dimensions)
# 获得总行数，总列数
print(sheet.max_row, sheet.max_column)

"""获取单元格的值"""
# 获取第一行第一列的元素，从1开始
print(sheet.cell(1, 1).value)
print(sheet['C3'].value)

"""循环遍历所有单元格"""
# # 法一
# for i in range(1, sheet.max_row + 1):
#     for j in range(1, sheet.max_column + 1):
#         print(f'{i}行{j}列： ', sheet.cell(i, j).value, end='\t')
#     print()
# # 法二
# for row in sheet.rows:
# # for row in sheet.iter_rows(min_row=2):  # 从第二行开始循环
#     for cell in row:
#         print(cell.value, end='@@@')
#     print()
# 法三 ---> 通过行列的字符拼接获得单元格位置。
for row_ch in range(1, 25):
    for col_ch in 'ABCDEFGHIJKLMN':
        print(sheet[f'{col_ch}{row_ch}'].value, end="\t")
    print()
